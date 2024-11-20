import openpyxl

def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_row)

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_column)

def readData(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum, column=columnno).value

def writeData(file,sheetName,rownum,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(file)

def add_fax_Detail_into_TestData(Totalcount,Sheetname,ColNameArray,row_num,Parameter):
    wb = openpyxl.load_workbook('C:\RobotFramework\Data\TestData.xlsx')
    sheet = wb.get_sheet_by_name(Sheetname)
    for x in range(Totalcount):
        sheet[ColNameArray[x] + str(row_num)].value = Parameter[x]
        wb.save('C:\RobotFramework\Data\TestData.xlsx')

def FindCoverage(list,Srchlist):
    existlist = list.count(Srchlist)
     
